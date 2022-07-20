from this import d
from bs4 import BeautifulSoup
import bs4
import requests
import openpyxl

wb = openpyxl.load_workbook ('D:\Python\APMC\APMC1.xlsx')
sheet = wb['GulTek']

#While loop

D = ('3101')
R = int(591)
  

r = requests.get('http://www.puneapmc.org/history.aspx?id=Rates'+D)
   #print (r)

SOUP = bs4.BeautifulSoup(r.text,features= 'html.parser')
    #print (SOUP)

    #Date
DATEX = SOUP.find_all('h2')[0].text
DTR = sheet.cell(row=R, column=1).value = DATEX

    #Onion
onionarr = SOUP.find_all('td')[3].text
onionmin = SOUP.find_all('td')[4].text
onionmax = SOUP.find_all('td')[5].text   

    #print (onionarr, onionmin,onionmax)
arr = sheet.cell(row=R, column=2).value = onionarr  
min = sheet.cell(row=R, column=3).value = onionmin
max = sheet.cell(row=R, column=4).value = onionmax

    #Potato
potatoarr = SOUP.find_all('td')[9].text
potatomin = SOUP.find_all('td')[10].text
potatomax = SOUP.find_all('td')[11].text  

    #print (potatoarr, potatonmin,potatonmax)
arr = sheet.cell(row=R, column=5).value = potatoarr
min = sheet.cell(row=R, column=6).value = potatomin
max = sheet.cell(row=R, column=7).value = potatomax

    #Garlic
garlicarr = SOUP.find_all('td')[15].text
garlicmin = SOUP.find_all('td')[16].text
garlicmax = SOUP.find_all('td')[17].text  
    #print (onionarr, onionmin,onionmax)

arr = sheet.cell(row=R, column=8).value = garlicarr
min = sheet.cell(row=R, column=9).value = garlicmin
max = sheet.cell(row=R, column=10).value = garlicmax

    #Ginger
gingerarr = SOUP.find_all('td')[3].text
gingermin = SOUP.find_all('td')[4].text 
gingermax = SOUP.find_all('td')[5].text  
    
    #print (onionarr, onionmin,onionmax)
arr = sheet.cell(row=R, column=11).value = gingerarr
min = sheet.cell(row=R, column=12).value = gingermin
max = sheet.cell(row=R, column=13).value = gingermax

wb.save('D:\Python\APMC\APMC1.xlsx')







  

  